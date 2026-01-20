# -*- coding: utf-8 -*-
"""
Created on Wed Dec 24 17:34:51 2025

@author: vvol-80
"""


import os
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import seaborn as sns
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import numpy as np
from scipy.stats import pearsonr
from datetime import timedelta
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from scipy.stats import linregress
import warnings
warnings.filterwarnings("ignore")  # чтобы не засорять лог предупреждениями statsmodels

# Импорт календаря
try:
    from tkcalendar import DateEntry
    HAS_TKCALENDAR = True
except ImportError:
    HAS_TKCALENDAR = False

# Попытка импорта squarify (опциональная библиотека)
try:
    import squarify
    HAS_SQUARIFY = True
except ImportError:
    HAS_SQUARIFY = False

sns.set(style="whitegrid")

# ===== DASHBOARD THEME =====
DASH_ORANGE = "#ff8c1a"
DASH_GRAY = "#f5f5f5"
DASH_BORDER = "#dddddd"
DASH_GREEN = "#2ecc71"
DASH_RED = "#e74c3c"

plt.rcParams.update({
    "axes.edgecolor": DASH_BORDER,
    "axes.labelcolor": "#333333",
    "xtick.color": "#666666",
    "ytick.color": "#666666",
    "axes.titleweight": "bold"
})

plt.rcParams.update({'font.size': 10})


class SalesAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sales Analyzer Pro v7.42 (Enhanced Analysis)")
        self.root.geometry("1200x900")
        self.root.minsize(1000, 700)
        self.df = None
        self.file_path = None
        self.results_storage = {}
        self.chk_vars = {}

        # Главный notebook (вкладки)
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True)

        # Вкладка настроек и анализа
        self.frame_main = ttk.Frame(self.notebook)
        self.notebook.add(self.frame_main, text="Настройки и анализ")

        # Вкладка дашборда
        self.frame_dashboard = ttk.Frame(self.notebook)
        self.notebook.add(self.frame_dashboard, text="Дашборд 📊")

        self._build_main_tab()
        self.canvas_dict = {}
        
        self.dashboard_cache = {}
        self.last_dashboard_key = None
        
        self.dashboard_nb = None
        self.date_col = None
        self.cat_col = None
        self.lbl_col = None
        self.df_filtered = None
        self.filter_date_from = None
        self.filter_date_to = None
        self.filter_product = None
        self.category_listbox = None
        
                # Графики топ/bottom в основном окне
        plt.rcParams.update({
            "axes.edgecolor": DASH_BORDER,
            "axes.labelcolor": "#333333",
            "xtick.color": "#666666",
            "ytick.color": "#666666",
            "axes.titleweight": "bold",
            "font.size": 7,  # Базовый размер шрифта
            "axes.titlesize": 9,  # Размер заголовков осей
        })

    def _build_main_tab(self):
        # ===== SCROLLABLE FRAME =====
        main_frame = tk.Frame(self.frame_main)
        main_frame.pack(fill="both", expand=True)
        
        # ===== ДВЕ ПАНЕЛИ (фиксированная ширина справа) =====
        content = ttk.Frame(main_frame)
        content.pack(fill="both", expand=True)
        
        left_panel = ttk.Frame(content)
        left_panel.pack(side="left", fill="both", expand=True)
        
        right_panel = ttk.Frame(content)
        right_panel.pack(side="right", fill="both")
        
        right_panel.configure(width=1300)
        right_panel.pack_propagate(False)
        
        self.dashboard_side = right_panel



        canvas = tk.Canvas(left_panel)
        scrollbar = ttk.Scrollbar(left_panel, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)


        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # ===== BLOCK 1: Загрузка файла =====
        frame_file = ttk.LabelFrame(scrollable_frame, text="1. Загрузка и проверка", padding=10)
        frame_file.pack(fill="x", padx=10, pady=5)

        btn_frame = ttk.Frame(frame_file)
        btn_frame.pack(fill="x")
        
        # Кнопка загрузки файла
        btn_load = ttk.Button(
            btn_frame,
            text="📁 Выбрать файл",
            command=self.load_file,
            width=15
        )
        btn_load.pack(side="left", padx=5)
        
        self.btn_check = ttk.Button(
            btn_frame,
            text="Проверка данных",
            command=self.open_data_check,
            state="disabled"
        )
        self.btn_check.pack(side="left", padx=5)

        self.lbl_status = ttk.Label(frame_file, text="Файл не выбран", foreground="gray")
        self.lbl_status.pack(anchor="w", pady=5)

        # ===== BLOCK 2: Настройка колонок =====
        frame_cols = ttk.LabelFrame(scrollable_frame, text="2. Настройка колонок", padding=10)
        frame_cols.pack(fill="x", padx=10, pady=5)

        self.combo_cat = self._combo(frame_cols, "Категория 1 *", 0)
        self.combo_cat2 = self._combo(frame_cols, "Категория 2 (Регион/Менеджер)", 1) # <--- Новое
        self.combo_label = self._combo(frame_cols, "Продукт *", 2) # Сдвигаем row
        self.combo_date = self._combo(frame_cols, "Дата (опц.)", 3)
        self.combo_cost = self._combo(frame_cols, "Себестоимость (опц.)", 4)

        self.calc_mode = tk.StringVar(value="sum")
        ttk.Separator(frame_cols, orient="horizontal").grid(row=5, columnspan=2, sticky="ew", pady=10) # row изменился на 5

        ttk.Radiobutton(frame_cols, text="Есть колонка Сумма",
                        variable=self.calc_mode, value="sum",
                        command=self.toggle_inputs).grid(row=6, column=0, sticky="w", pady=4) # row 6
        ttk.Radiobutton(frame_cols, text="Цена × Кол-во",
                        variable=self.calc_mode, value="calc",
                        command=self.toggle_inputs).grid(row=6, column=1, sticky="w", pady=4) # row 6

        self.lbl_param1 = ttk.Label(frame_cols, text="Сумма продаж:")
        self.lbl_param1.grid(row=7, column=0, sticky="w", pady=4) # row 7
        self.combo_param1 = ttk.Combobox(frame_cols, state="readonly", width=30)
        self.combo_param1.grid(row=7, column=1, padx=5, pady=4) # row 7
        
        self.lbl_param2 = ttk.Label(frame_cols, text="Кол-во:")
        self.combo_param2 = ttk.Combobox(frame_cols, state="disabled", width=30)

        # ===== BLOCK 3: Выбор анализов =====
        frame_options = ttk.LabelFrame(scrollable_frame, text="3. Выберите анализы", padding=10)
        frame_options.pack(fill="x", padx=10, pady=25)

        options = [
            ("finance", "Расчёт прибыли и маржи"),
            ("share", "Доли в выручке"),
            ("var_coeff", "Коэффициент вариации"),
            ("top_bot_sales", "TOP-5 / BOTTOM-5 по продажам"),
            ("top_bot_profit", "TOP-5 по прибыли"),
            ("top_bot_margin", "TOP-5 по марже"),
            ("abc_xyz", "ABC-XYZ анализ по товарам"),
            ("abc_xyz_cat", "ABC-XYZ анализ по категориям"),
            ("time_trends", "Динамика по времени"),
            ("seasonality", "Месячный анализ сезонности"),
            ("elasticity", "Эластичность спроса"),
            ("correlation", "Корреляция цена-колво"),
            ("forecast", "Прогноз продаж на 30 дней вперед"),
        ]

        for key, text in options:
            var = tk.BooleanVar(value=True if key in ["finance", "share", "var_coeff", "top_bot_sales", "top_bot_profit", "top_bot_margin",
                                                      "abc_xyz", "abc_xyz_cat", "time_trends", "seasonality", "elasticity", "correlation", "forecast"] else False)
            self.chk_vars[key] = var
            ttk.Checkbutton(frame_options, text=text, variable=var).pack(anchor="w", pady=2)

        # ===== BLOCK ACTIONS =====
        frame_actions = ttk.Frame(scrollable_frame, padding=10)
        frame_actions.pack(fill="x", pady=10)

        self.btn_analyze = ttk.Button(frame_actions, text="АНАЛИЗИРОВАТЬ 🚀",
                                      command=self.run_analysis, state="disabled")
        self.btn_analyze.pack(side="left", expand=True, fill="x", padx=5)

        self.btn_export = ttk.Button(frame_actions, text="Экспорт в Excel",
                                     command=self.export_excel, state="disabled")
        self.btn_export.pack(side="left", padx=5)

        # ===== LOG =====
        frame_log = ttk.LabelFrame(scrollable_frame, text="Лог выполнения", padding=10)
        frame_log.pack(fill="both", expand=True, padx=10, pady=5)

        self.txt_log = tk.Text(frame_log, height=12, font=("Consolas", 10), wrap="word")
        self.txt_log.pack(fill="both", expand=True)

    def _combo(self, parent, text, row):
        ttk.Label(parent, text=text).grid(row=row, column=0, sticky="w", pady=4)
        cb = ttk.Combobox(parent, state="readonly", width=30)
        cb.grid(row=row, column=1, padx=5, pady=4)
        return cb

    def toggle_inputs(self):
            """Переключение между режимами расчета: Сумма или Цена х Кол-во"""
            mode = self.calc_mode.get()
            if mode == "sum":
                self.lbl_param1.config(text="Сумма продаж:")
                self.lbl_param2.config(text="Прибыль (опц.):")
                
                # Принудительно задаем правильные строки, чтобы ничего не накладывалось
                self.lbl_param1.grid(row=7, column=0, sticky="w", pady=4)
                self.combo_param1.grid(row=7, column=1, padx=5, pady=4)
                
                self.lbl_param2.grid(row=8, column=0, sticky="w", pady=4)
                self.combo_param2.grid(row=8, column=1, padx=5, pady=4)
                
                self.combo_param2.config(state="readonly")
            else:
                self.lbl_param1.config(text="Цена за единицу:")
                self.lbl_param2.config(text="Количество:")
                
                # Здесь тоже ставим 7 и 8
                self.lbl_param1.grid(row=7, column=0, sticky="w", pady=4)
                self.combo_param1.grid(row=7, column=1, padx=5, pady=4)
                
                self.lbl_param2.grid(row=8, column=0, sticky="w", pady=4)
                self.combo_param2.grid(row=8, column=1, padx=5, pady=4)
                
                self.combo_param2.config(state="readonly")
                
    def log(self, message):
        self.txt_log.insert(tk.END, message + "\n")
        self.txt_log.see(tk.END)
        self.root.update_idletasks()

    # ================= HELPERS (NEW) =================
    def detect_type(self, series):
        """Определяет тип данных в колонке для отчета."""
        if pd.api.types.is_numeric_dtype(series):
            return "Число"
        elif pd.api.types.is_datetime64_any_dtype(series):
            return "Дата"
        else:
            return "Строка"

    # ================= FILE & DATA QUALITY =================
    def load_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel и CSV", "*.xlsx *.xls *.csv")]
        )
        if not path:
            return
        try:
            if path.endswith(('.xlsx', '.xls')):
                self.df = pd.read_excel(path)
            else:
                self.df = pd.read_csv(path, encoding='utf-8', on_bad_lines='skip')

            self.file_path = path
            cols = list(self.df.columns)
            for cb in [self.combo_cat, self.combo_cat2, self.combo_label, self.combo_date,
                       self.combo_cost, self.combo_param1, self.combo_param2]:
                cb['values'] = cols
                cb.set("")

            self.lbl_status.config(text=f"{os.path.basename(path)} | {len(self.df):,} строк", foreground="green")
            self.btn_check.config(state="normal")
            self.btn_analyze.config(state="normal")
            if self.df.isnull().values.any():
                messagebox.showinfo("Предупреждение", "Есть ошибки данных. Рекомендуется проверка качества.")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    # ================= DATA QUALITY =================
    def open_data_check(self):
        if self.df is None: return
        win = tk.Toplevel(self.root)
        win.title("Диагностика и Исправление")
        win.geometry("950x650")
        win.grab_set()

        # Layout
        win.rowconfigure(0, weight=1)
        win.rowconfigure(1, weight=0)
        win.columnconfigure(0, weight=1)

        # 1. Отчет
        txt_frame = ttk.Frame(win)
        txt_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        txt = tk.Text(txt_frame, font=("Consolas", 10), wrap="none")
        vsb = ttk.Scrollbar(txt_frame, command=txt.yview); hsb = ttk.Scrollbar(txt_frame, orient="horizontal", command=txt.xview)
        txt.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x"); txt.pack(side="left", fill="both", expand=True)

        # 2. Панель исправления
        fix_panel = ttk.LabelFrame(win, text="Инструменты исправления данных", padding=10)
        fix_panel.grid(row=1, column=0, sticky="ew", padx=10, pady=10)

        # Переменные для выбора режима
        nan_mode = tk.StringVar(value="zero")
        neg_mode = tk.StringVar(value="abs")

        # Визуальный блок для NaN
        f_nan = ttk.Frame(fix_panel)
        f_nan.pack(side="left", padx=20)
        ttk.Label(f_nan, text="Пустые (NaN):", font=("Arial", 10, "bold")).pack(anchor="w")
        ttk.Radiobutton(f_nan, text="Заменить на 0", variable=nan_mode, value="zero").pack(anchor="w")
        ttk.Radiobutton(f_nan, text="Среднее значение", variable=nan_mode, value="mean").pack(anchor="w")
        ttk.Radiobutton(f_nan, text="Удалить строки", variable=nan_mode, value="drop").pack(anchor="w")

        # Визуальный блок для Отрицательных
        f_neg = ttk.Frame(fix_panel)
        f_neg.pack(side="left", padx=20)
        ttk.Label(f_neg, text="Отрицательные:", font=("Arial", 10, "bold")).pack(anchor="w")
        ttk.Radiobutton(f_neg, text="В модуль (сделать +)", variable=neg_mode, value="abs").pack(anchor="w")
        ttk.Radiobutton(f_neg, text="Заменить на 0", variable=neg_mode, value="zero").pack(anchor="w")

        def run_fix():
            # Исправляем NaN
            num_cols = self.df.select_dtypes(include=[np.number]).columns
            if nan_mode.get() == "zero":
                self.df.fillna(0, inplace=True)
            elif nan_mode.get() == "mean":
                for c in num_cols: self.df[c].fillna(self.df[c].mean(), inplace=True)
            elif nan_mode.get() == "drop":
                self.df.dropna(inplace=True)

            # Исправляем Отрицательные во ВСЕХ числовых колонках
            fix_count = 0
            for c in num_cols:
                neg_mask = self.df[c] < 0
                count = neg_mask.sum()
                if count > 0:
                    if neg_mode.get() == "abs":
                        self.df[c] = self.df[c].abs()
                    else:
                        self.df.loc[neg_mask, c] = 0
                    fix_count += count
            
            messagebox.showinfo("Готово", f"Данные очищены!\nИсправлено отрицательных значений: {fix_count}")
            self.lbl_status.config(text=f"Очищено | {len(self.df)} строк", foreground="blue")
            win.destroy()

        ttk.Button(fix_panel, text="🚀 ПРИМЕНИТЬ ИСПРАВЛЕНИЯ", command=run_fix, width=30).pack(side="right", padx=10)

        # Заполнение отчета
        txt.insert("1.0", f"{'Колонка':<30} | {'Тип':<10} | {'NaN':<7} | {'Отриц.':<7} | {'Статус'}\n")
        txt.insert(tk.END, "-"*80 + "\n")
        for col in self.df.columns:
            s = self.df[col]
            nan_c = s.isna().sum()
            neg_c = (s < 0).sum() if pd.api.types.is_numeric_dtype(s) else 0
            status = "❌ ОШИБКИ" if (nan_c > 0 or neg_c > 0) else "✅ ОК"
            txt.insert(tk.END, f"{str(col)[:28]:<30} | {str(s.dtype):<10} | {nan_c:<7} | {neg_c:<7} | {status}\n")

    # ================= ENHANCED ANALYSIS =================
    def run_analysis(self):
        if self.df is None:
            messagebox.showwarning("Нет данных", "Загрузите файл.")
            return
    
        if self.calc_mode.get() == "sum":
            if not self.combo_param1.get():
                messagebox.showwarning("Ошибка", "Выберите колонку с суммой продаж")
                return
        else:
            if not self.combo_param1.get() or not self.combo_param2.get():
                messagebox.showwarning("Ошибка", "Выберите колонки Цены и Количества")
                return
        self.df_base = self.df.copy()
        self.txt_log.delete(1.0, tk.END)
        self.log("🚀 Запуск расширенного анализа...")
        self.results_storage = {}
        
        # FIX: Инициализируем переменные, чтобы избежать UnboundLocalError
        prod_stats = pd.DataFrame()
        cat_stats = pd.DataFrame()
    
        cat_col = self.combo_cat.get() or None
        cat2_col = self.combo_cat2.get() or None
        lbl_col = self.combo_label.get() or None
        cost_col = self.combo_cost.get() or None
        date_col = self.combo_date.get() or None
    
        # Расчёт продаж
        if self.calc_mode.get() == "sum":
            sales_col = self.combo_param1.get()
            self.df['Продажи'] = pd.to_numeric(self.df[sales_col], errors='coerce').fillna(0)
            self.log(f"📊 Использована колонка суммы: {sales_col}")
        else:
            price_col = self.combo_param1.get()
            qty_col = self.combo_param2.get()
            self.df['_Price'] = pd.to_numeric(self.df[price_col], errors='coerce').fillna(0)
            self.df['_Qty'] = pd.to_numeric(self.df[qty_col], errors='coerce').fillna(0)
            self.df['Продажи'] = self.df['_Price'] * self.df['_Qty']
            self.log(f"📊 Продажи рассчитаны: {price_col} × {qty_col}")
    
        total_sales = self.df['Продажи'].sum()
        self.log(f"💰 Общий объём продаж: {total_sales:,.2f}")
    
        # Прибыль и маржа
        has_profit = False
        if cost_col and self.chk_vars['finance'].get():
            cost_series = pd.to_numeric(self.df[cost_col], errors='coerce').fillna(0)
            if self.calc_mode.get() == "calc":
                self.df['Себестоимость_Общ'] = cost_series * self.df['_Qty']
            else:
                self.df['Себестоимость_Общ'] = cost_series
            self.df['Прибыль'] = self.df['Продажи'] - self.df['Себестоимость_Общ']
            self.df['Маржа_%'] = np.where(self.df['Продажи'] == 0, 0, self.df['Прибыль'] / self.df['Продажи'] * 100)
            has_profit = True
            self.log("💵 Рассчитана прибыль и маржа")

        # === Метрики по продуктам ===
        if lbl_col:
            prod_stats = self.df.groupby(lbl_col).agg(
                Суммарные_Продажи=('Продажи', 'sum'),
                Средние_Продажи=('Продажи', 'mean'),
                Медиана=('Продажи', 'median'),
                StdDev=('Продажи', 'std'),
                Количество_транзакций=('Продажи', 'count')
            ).reset_index()

            if has_profit:
                profit_prod = self.df.groupby(lbl_col).agg(
                    Суммарная_Прибыль=('Прибыль', 'sum'),
                    Средняя_Маржа=('Маржа_%', 'mean')
                ).reset_index()
                prod_stats = prod_stats.merge(profit_prod, on=lbl_col, how='left')

            if self.chk_vars['share'].get():
                prod_stats['Доля_%'] = prod_stats['Суммарные_Продажи'] / total_sales * 100 if total_sales else 0

            # Коэффициент вариации
            if self.chk_vars['var_coeff'].get():
                # Проверяем, что StdDev есть в DataFrame
                if 'StdDev' in prod_stats.columns:
                    prod_stats['Коэф_вариации_%'] = np.where(
                        prod_stats['Средние_Продажи'] == 0, 
                        0, 
                        (prod_stats['StdDev'] / prod_stats['Средние_Продажи']) * 100
                    )
                    prod_stats['Стабильность'] = prod_stats['Коэф_вариации_%'].apply(
                        lambda x: 'Высокая' if x <= 30 else 'Средняя' if x <= 60 else 'Низкая'
                    )
                    self.log("📈 Рассчитан коэффициент вариации по продуктам")
                else:
                    self.log("⚠ Невозможно рассчитать коэффициент вариации: отсутствует StdDev")

            self.results_storage['Метрики_по_продуктам'] = prod_stats
        else:
            self.log("ℹ Колонка 'Продукт' не выбрана — анализ по продуктам пропущен")

        # === Метрики по категориям ===
        if cat_col:
            # Сначала собираем базовую статистику
            cat_stats = self.df.groupby(cat_col).agg(
                Суммарные_Продажи=('Продажи', 'sum'),
                Средние_Продажи=('Продажи', 'mean'),
                Количество_транзакций=('Продажи', 'count')
            ).reset_index()

            # Добавляем StdDev отдельно
            cat_std = self.df.groupby(cat_col)['Продажи'].std().reset_index(name='StdDev')
            cat_stats = cat_stats.merge(cat_std, on=cat_col, how='left')

            if has_profit:
                profit_cat = self.df.groupby(cat_col).agg(
                    Суммарная_Прибыль=('Прибыль', 'sum')
                ).reset_index()
                cat_stats = cat_stats.merge(profit_cat, on=cat_col, how='left')

            if self.chk_vars['share'].get():
                cat_stats['Доля_%'] = cat_stats['Суммарные_Продажи'] / total_sales * 100 if total_sales else 0

            # Коэффициент вариации для категорий
            if self.chk_vars['var_coeff'].get():
                # Проверяем, что StdDev есть в DataFrame
                if 'StdDev' in cat_stats.columns:
                    cat_stats['Коэф_вариации_%'] = np.where(
                        cat_stats['Средние_Продажи'] == 0, 
                        0, 
                        (cat_stats['StdDev'] / cat_stats['Средние_Продажи']) * 100
                    )
                    cat_stats['Стабильность'] = cat_stats['Коэф_вариации_%'].apply(
                        lambda x: 'Высокая' if x <= 30 else 'Средняя' if x <= 60 else 'Низкая'
                    )
                else:
                    self.log("⚠ Невозможно рассчитать коэффициент вариации для категорий: отсутствует StdDev")

            self.results_storage['Метрики_по_категориям'] = cat_stats
        else:
            self.log("ℹ Колонка 'Категория' не выбрана — анализ по категориям пропущен")
            
           # === Метрики по Категории 2 (Менеджер/Регион) ===
        if cat2_col:
            cat2_stats = self.df.groupby(cat2_col).agg(
                Суммарные_Продажи=('Продажи', 'sum'),
                Средние_Продажи=('Продажи', 'mean'),
                Количество_транзакций=('Продажи', 'count')
            ).reset_index()

            cat2_std = self.df.groupby(cat2_col)['Продажи'].std().reset_index(name='StdDev')
            cat2_stats = cat2_stats.merge(cat2_std, on=cat2_col, how='left')

            if has_profit:
                profit_cat2 = self.df.groupby(cat2_col).agg(
                    Суммарная_Прибыль=('Прибыль', 'sum')
                ).reset_index()
                cat2_stats = cat2_stats.merge(profit_cat2, on=cat2_col, how='left')

            if self.chk_vars['share'].get():
                cat2_stats['Доля_%'] = cat2_stats['Суммарные_Продажи'] / total_sales * 100 if total_sales else 0

            self.results_storage['Метрики_по_категории_2'] = cat2_stats
            self.log(f"📊 Рассчитана статистика по '{cat2_col}'")
        # === Общие метрики ===
        overall = pd.DataFrame({
            'Показатель': ['Общая выручка', 'Количество транзакций', 'Средняя продажа'],
            'Значение': [total_sales, len(self.df), self.df['Продажи'].mean() if len(self.df) > 0 else 0]
        })
        if has_profit:
            overall = pd.concat([overall, pd.DataFrame({
                'Показатель': ['Общая прибыль', 'Средняя маржа %'],
                'Значение': [self.df['Прибыль'].sum(), self.df['Маржа_%'].mean()]
            })], ignore_index=True)
        
        # Добавляем коэффициент вариации в общие метрики
        if self.chk_vars['var_coeff'].get():
            overall_coeff = pd.DataFrame({
                'Показатель': ['Коэффициент вариации (общий)'],
                'Значение': [self.df['Продажи'].std() / self.df['Продажи'].mean() * 100 if self.df['Продажи'].mean() > 0 else 0]
            })
            overall = pd.concat([overall, overall_coeff], ignore_index=True)
        
        self.results_storage['Общие_показатели'] = overall

        # Сохраняем основные таблицы (если они не пустые)
        if not prod_stats.empty:
            self.results_storage['Метрики_по_продуктам'] = prod_stats
        if not cat_stats.empty:
            self.results_storage['Метрики_по_категориям'] = cat_stats

        # ===== TOP/BOTTOM =====
        # Продукты
        if not prod_stats.empty:
            if self.chk_vars['top_bot_sales'].get():
                cols = [lbl_col, 'Суммарные_Продажи']
                if 'Доля_%' in prod_stats.columns: cols.append('Доля_%')
                if 'Коэф_вариации_%' in prod_stats.columns: cols.append('Коэф_вариации_%')
                self.results_storage['TOP_5_Продукты_Продажи'] = prod_stats.nlargest(5, 'Суммарные_Продажи')[cols]
                self.results_storage['BOTTOM_5_Продукты_Продажи'] = prod_stats.nsmallest(5, 'Суммарные_Продажи')[cols]

            if self.chk_vars['top_bot_profit'].get() and has_profit:
                self.results_storage['TOP_5_Продукты_Прибыль'] = prod_stats.nlargest(5, 'Суммарная_Прибыль')[[lbl_col, 'Суммарная_Прибыль', 'Средняя_Маржа']]

        # Категории
        if not cat_stats.empty:
            if self.chk_vars['top_bot_sales'].get():
                cols = [cat_col, 'Суммарные_Продажи']
                if 'Доля_%' in cat_stats.columns: cols.append('Доля_%')
                if 'Коэф_вариации_%' in cat_stats.columns: cols.append('Коэф_вариации_%')
                self.results_storage['TOP_5_Категории_Продажи'] = cat_stats.nlargest(5, 'Суммарные_Продажи')[cols]
            
            if self.chk_vars['top_bot_profit'].get() and has_profit:
                self.results_storage['TOP_5_Категории_Прибыль'] = cat_stats.nlargest(5, 'Суммарная_Прибыль')[[cat_col, 'Суммарная_Прибыль']]

        # ===== ABC-XYZ анализ по товарам =====
        if self.chk_vars['abc_xyz'].get() and not prod_stats.empty and 'StdDev' in prod_stats.columns:
            abc_prod = prod_stats.sort_values('Суммарные_Продажи', ascending=False).copy()
            if total_sales > 0:
                abc_prod['CumShare'] = abc_prod['Суммарные_Продажи'].cumsum() / total_sales
                abc_prod['ABC'] = abc_prod['CumShare'].apply(lambda x: 'A' if x <= 0.80 else 'B' if x <= 0.95 else 'C')
                abc_prod['Coeff_Var'] = np.where(abc_prod['Средние_Продажи'] == 0, 0, abc_prod['StdDev'] / abc_prod['Средние_Продажи'])
                abc_prod['XYZ'] = abc_prod['Coeff_Var'].apply(lambda x: 'X' if x <= 0.25 else 'Y' if x <= 0.50 else 'Z')
                abc_prod['ABC_XYZ'] = abc_prod['ABC'] + abc_prod['XYZ']
                
                # Добавляем интерпретацию
                abc_xyz_map = {
                    'AX': 'Ключевые стабильные',
                    'AY': 'Ключевые умеренные',
                    'AZ': 'Ключевые нестабильные',
                    'BX': 'Второстепенные стабильные',
                    'BY': 'Второстепенные умеренные',
                    'BZ': 'Второстепенные нестабильные',
                    'CX': 'Мелкие стабильные',
                    'CY': 'Мелкие умеренные',
                    'CZ': 'Мелкие нестабильные'
                }
                abc_prod['Интерпретация'] = abc_prod['ABC_XYZ'].map(abc_xyz_map)
                
                # Статистика по группам
                group_stats = abc_prod.groupby('ABC_XYZ').agg(
                    Кол_во_товаров=('ABC_XYZ', 'count'),
                    Суммарные_Продажи=('Суммарные_Продажи', 'sum'),
                    Доля_в_выручке=('Суммарные_Продажи', lambda x: x.sum() / total_sales * 100)
                ).reset_index()
                group_stats['Интерпретация'] = group_stats['ABC_XYZ'].map(abc_xyz_map)
                
                self.results_storage['ABC_XYZ_Продукты_Детально'] = abc_prod
                self.results_storage['ABC_XYZ_Продукты_Статистика'] = group_stats
                self.log("✅ ABC-XYZ анализ по товарам выполнен")
            else:
                self.log("⚠ Не удалось выполнить ABC-XYZ анализ: общие продажи равны 0")
        elif self.chk_vars['abc_xyz'].get() and not prod_stats.empty:
            self.log("⚠ Не удалось выполнить ABC-XYZ анализ: отсутствует StdDev")

        # ===== ABC-XYZ анализ по категориям =====
        if self.chk_vars.get('abc_xyz_cat', tk.BooleanVar()).get() and not cat_stats.empty and 'StdDev' in cat_stats.columns:
            abc_cat = cat_stats.sort_values('Суммарные_Продажи', ascending=False).copy()
            if total_sales > 0:
                abc_cat['CumShare'] = abc_cat['Суммарные_Продажи'].cumsum() / total_sales
                abc_cat['ABC'] = abc_cat['CumShare'].apply(lambda x: 'A' if x <= 0.80 else 'B' if x <= 0.95 else 'C')
                
                # Используем уже существующий StdDev
                abc_cat['Coeff_Var'] = np.where(
                    abc_cat['Средние_Продажи'] == 0, 
                    0, 
                    abc_cat['StdDev'] / abc_cat['Средние_Продажи']
                )
                abc_cat['XYZ'] = abc_cat['Coeff_Var'].apply(
                    lambda x: 'X' if x <= 0.10 else 'Y' if x <= 0.25 else 'Z'
                )
                abc_cat['ABC_XYZ'] = abc_cat['ABC'] + abc_cat['XYZ']
                
                self.results_storage['ABC_XYZ_Категории'] = abc_cat
                self.log("✅ ABC-XYZ анализ по категориям выполнен")
            else:
                self.log("⚠ Не удалось выполнить ABC-XYZ анализ по категориям: общие продажи равны 0")
        elif self.chk_vars.get('abc_xyz_cat', tk.BooleanVar()).get() and not cat_stats.empty:
            self.log("⚠ Не удалось выполнить ABC-XYZ анализ по категориям: отсутствует StdDev")

        # ===== Динамика по времени =====
        # ... остальной код остается без изменений ...

        # ===== Динамика по времени =====
        if self.chk_vars['time_trends'].get() and date_col:
            try:
                self.df[date_col] = pd.to_datetime(self.df[date_col], errors='coerce')
                # FIX: Используем freq='M' для большей совместимости, 'ME' - для новых версий Pandas
                ts_month = self.df.groupby(pd.Grouper(key=date_col, freq='M'))['Продажи'].sum().reset_index()
                ts_month['Темп_роста_%'] = ts_month['Продажи'].pct_change() * 100
                ts_month['SMA_3'] = ts_month['Продажи'].rolling(3, min_periods=1).mean()
                self.results_storage['Динамика_по_месяцам'] = ts_month
                
                self.df['Год'] = self.df[date_col].dt.year
                ts_year = self.df.groupby('Год')['Продажи'].sum().reset_index()
                self.results_storage['Динамика_по_годам'] = ts_year
            except Exception as e:
                self.log(f"Ошибка при расчёте динамики: {e}")

        # ===== МЕСЯЧНЫЙ АНАЛИЗ СЕЗОННОСТИ =====
        if self.chk_vars.get('seasonality', tk.BooleanVar()).get() and date_col:
            try:
                self.df[date_col] = pd.to_datetime(self.df[date_col], errors='coerce')
                
                # Анализ по месяцам
                self.df['Месяц'] = self.df[date_col].dt.month
                self.df['Месяц_название'] = self.df[date_col].dt.strftime('%B')
                self.df['Год'] = self.df[date_col].dt.year
                
                # Месячная статистика
                month_stats = self.df.groupby(['Год', 'Месяц', 'Месяц_название']).agg(
                    Продажи=('Продажи', 'sum'),
                    Количество_транзакций=('Продажи', 'count'),
                    Средний_чек=('Продажи', 'mean')
                ).reset_index()
                
                # Сводная таблица по месяцам
                pivot_month = month_stats.pivot_table(
                    index='Месяц_название',
                    values='Продажи',
                    aggfunc='mean'
                ).reset_index()
                pivot_month.columns = ['Месяц', 'Средние_продажи_за_месяц']
                pivot_month = pivot_month.sort_values('Средние_продажи_за_месяц', ascending=False)
                
                # Сезонный индекс (относительно среднего)
                total_avg = pivot_month['Средние_продажи_за_месяц'].mean()
                pivot_month['Сезонный_индекс'] = pivot_month['Средние_продажи_за_месяц'] / total_avg
                pivot_month['Тип_сезона'] = pivot_month['Сезонный_индекс'].apply(
                    lambda x: 'Высокий' if x > 1.2 else 'Низкий' if x < 0.8 else 'Средний'
                )
                
                # Тренд по годам
                year_month_stats = self.df.groupby(['Год', 'Месяц']).agg({'Продажи': 'sum'}).reset_index()
                year_month_pivot = year_month_stats.pivot_table(
                    index='Месяц',
                    columns='Год',
                    values='Продажи',
                    aggfunc='sum'
                )
                
                self.results_storage['Сезонность_по_месяцам'] = pivot_month
                self.results_storage['Сезонность_детально'] = month_stats
                if not year_month_pivot.empty:
                    self.results_storage['Сезонность_по_годам_и_месяцам'] = year_month_pivot.reset_index()
                
                self.log("🌞 Анализ сезонности выполнен")
                
            except Exception as e:
                self.log(f"Ошибка при анализе сезонности: {e}")

        # ===== ЭЛАСТИЧНОСТЬ СПРОСА =====
        if self.chk_vars.get('elasticity', tk.BooleanVar()).get() and self.calc_mode.get() == "calc":
            try:
                self.log("📊 Расчёт эластичности спроса...")
                
                elasticity_results = []
                
                if lbl_col:
                    # По продуктам
                    for product in self.df[lbl_col].unique():
                        product_df = self.df[self.df[lbl_col] == product]
                        if len(product_df) >= 5:  # Минимум 5 наблюдений
                            clean_df = product_df[['_Price', '_Qty']].dropna()
                            if len(clean_df) > 1:
                                try:
                                    # Логарифмическая регрессия для эластичности
                                    log_price = np.log(clean_df['_Price'])
                                    log_qty = np.log(clean_df['_Qty'])
                                    
                                    slope, intercept, r_value, p_value, std_err = linregress(log_price, log_qty)
                                    elasticity = slope  # Коэффициент эластичности
                                    
                                    elasticity_results.append({
                                        'Тип': 'Продукт',
                                        'Название': product,
                                        'Эластичность': elasticity,
                                        'R_квадрат': r_value**2,
                                        'p_value': p_value,
                                        'Интерпретация': 'Эластичный' if elasticity < -1 else 'Неэластичный' if elasticity > -1 else 'Единичная'
                                    })
                                except:
                                    continue
                
                if cat_col:
                    # По категориям
                    for category in self.df[cat_col].unique():
                        cat_df = self.df[self.df[cat_col] == category]
                        if len(cat_df) >= 5:
                            clean_df = cat_df[['_Price', '_Qty']].dropna()
                            if len(clean_df) > 1:
                                try:
                                    log_price = np.log(clean_df['_Price'])
                                    log_qty = np.log(clean_df['_Qty'])
                                    
                                    slope, intercept, r_value, p_value, std_err = linregress(log_price, log_qty)
                                    elasticity = slope
                                    
                                    elasticity_results.append({
                                        'Тип': 'Категория',
                                        'Название': category,
                                        'Эластичность': elasticity,
                                        'R_квадрат': r_value**2,
                                        'p_value': p_value,
                                        'Интерпретация': 'Эластичный' if elasticity < -1 else 'Неэластичный' if elasticity > -1 else 'Единичная'
                                    })
                                except:
                                    continue
                
                if elasticity_results:
                    elasticity_df = pd.DataFrame(elasticity_results)
                    elasticity_df = elasticity_df.sort_values('Эластичность')
                    
                    # Сводная статистика
                    elasticity_summary = pd.DataFrame({
                        'Показатель': [
                            'Средняя эластичность',
                            'Медианная эластичность',
                            'Доля эластичных товаров',
                            'Доля неэластичных товаров'
                        ],
                        'Значение': [
                            elasticity_df['Эластичность'].mean(),
                            elasticity_df['Эластичность'].median(),
                            len(elasticity_df[elasticity_df['Интерпретация'] == 'Эластичный']) / len(elasticity_df) * 100,
                            len(elasticity_df[elasticity_df['Интерпретация'] == 'Неэластичный']) / len(elasticity_df) * 100
                        ]
                    })
                    
                    self.results_storage['Эластичность_спроса_детально'] = elasticity_df
                    self.results_storage['Эластичность_спроса_сводка'] = elasticity_summary
                    self.log(f"✅ Эластичность спроса рассчитана для {len(elasticity_df)} позиций")
                
            except Exception as e:
                self.log(f"⚠ Ошибка при расчёте эластичности: {e}")

        # ===== КОРРЕЛЯЦИЯ ЦЕНА-КОЛИЧЕСТВО =====
        if self.chk_vars.get('correlation', tk.BooleanVar()).get() and self.calc_mode.get() == "calc":
            try:
                self.log("🔗 Анализ корреляции цена-количество...")
                
                correlation_results = []
                
                # Общая корреляция
                clean_df = self.df[['_Price', '_Qty']].dropna()
                if len(clean_df) > 1 and clean_df['_Price'].std() > 0 and clean_df['_Qty'].std() > 0:
                    corr, p = pearsonr(clean_df['_Price'], clean_df['_Qty'])
                    correlation_results.append({
                        'Уровень': 'Общий',
                        'Корреляция': corr,
                        'p_value': p,
                        'Интерпретация': 'Сильная отрицательная' if corr < -0.7 else 
                                        'Умеренная отрицательная' if corr < -0.3 else 
                                        'Слабая отрицательная' if corr < 0 else
                                        'Слабая положительная' if corr < 0.3 else
                                        'Умеренная положительная' if corr < 0.7 else 'Сильная положительная'
                    })
                
                if lbl_col:
                    # По продуктам
                    for product in self.df[lbl_col].unique()[:20]:  # Ограничиваем топ-20
                        product_df = self.df[self.df[lbl_col] == product]
                        clean_df = product_df[['_Price', '_Qty']].dropna()
                        if len(clean_df) > 5 and clean_df['_Price'].std() > 0 and clean_df['_Qty'].std() > 0:
                            corr, p = pearsonr(clean_df['_Price'], clean_df['_Qty'])
                            correlation_results.append({
                                'Уровень': 'Продукт',
                                'Название': product,
                                'Корреляция': corr,
                                'p_value': p,
                                'Интерпретация': 'Сильная отриц.' if corr < -0.7 else 
                                                'Умеренная отриц.' if corr < -0.3 else 
                                                'Слабая отриц.' if corr < 0 else
                                                'Слабая полож.' if corr < 0.3 else
                                                'Умеренная полож.' if corr < 0.7 else 'Сильная полож.'
                            })
                
                if cat_col:
                    # По категориям
                    for category in self.df[cat_col].unique():
                        cat_df = self.df[self.df[cat_col] == category]
                        clean_df = cat_df[['_Price', '_Qty']].dropna()
                        if len(clean_df) > 5 and clean_df['_Price'].std() > 0 and clean_df['_Qty'].std() > 0:
                            corr, p = pearsonr(clean_df['_Price'], clean_df['_Qty'])
                            correlation_results.append({
                                'Уровень': 'Категория',
                                'Название': category,
                                'Корреляция': corr,
                                'p_value': p,
                                'Интерпретация': 'Сильная отриц.' if corr < -0.7 else 
                                                'Умеренная отриц.' if corr < -0.3 else 
                                                'Слабая отриц.' if corr < 0 else
                                                'Слабая полож.' if corr < 0.3 else
                                                'Умеренная полож.' if corr < 0.7 else 'Сильная полож.'
                            })
                
                if correlation_results:
                    correlation_df = pd.DataFrame(correlation_results)
                    self.results_storage['Корреляция_цена_количество'] = correlation_df
                    self.log(f"✅ Корреляция рассчитана для {len(correlation_df)} позиций")
                
            except Exception as e:
                self.log(f"⚠ Ошибка при расчёте корреляции: {e}")

        # ===== ПРОГНОЗ ПРОДАЖ НА 30 ДНЕЙ =====
        if self.chk_vars.get('forecast', tk.BooleanVar()).get() and date_col:
            try:
                self.df[date_col] = pd.to_datetime(self.df[date_col], errors='coerce')
                self.df = self.df.dropna(subset=[date_col])  # удаляем строки без даты
                self.df = self.df.sort_values(date_col)

                forecast_days = 30
                last_date = self.df[date_col].max()
                future_dates = pd.date_range(start=last_date + timedelta(days=1), periods=forecast_days, freq='D')

                forecast_results = {}

                # Функция прогнозирования для одной группы
                def forecast_group(group_df, group_name, key_col):
                    if len(group_df) < 3:
                        return None  # слишком мало данных

                    ts_daily = group_df.set_index(date_col)['Продажи'].resample('D').sum().fillna(0)
                    if ts_daily.sum() == 0:
                        return None

                    # Попробуем Holt-Winters с сезонностью (недельной)
                    try:
                        model = ExponentialSmoothing(
                            ts_daily,
                            trend='add',
                            seasonal='add',
                            seasonal_periods=7  # недельная сезонность
                        ).fit()
                        forecast = model.forecast(forecast_days)
                    except:
                        try:
                            # Fallback: простое экспоненциальное сглаживание с трендом
                            model = ExponentialSmoothing(ts_daily, trend='add', seasonal=None).fit()
                            forecast = model.forecast(forecast_days)
                        except:
                            # Последний fallback: линейный тренд по последним 90 дням
                            recent = ts_daily[-90:]
                            if len(recent) > 1:
                                x = np.arange(len(recent))
                                coeffs = np.polyfit(x, recent.values, 1)
                                last_val = recent.iloc[-1]
                                trend_daily = coeffs[0]
                                forecast = [max(0, last_val + trend_daily * (i+1)) for i in range(forecast_days)]
                                forecast = pd.Series(forecast, index=future_dates)
                            else:
                                forecast = pd.Series([ts_daily.mean()] * forecast_days, index=future_dates)

                    forecast_df = pd.DataFrame({
                        'Дата': future_dates,
                        'Прогноз_Продажи': forecast.round(2)
                    })
                    forecast_df.insert(0, key_col, group_name)
                    return forecast_df

                # Прогноз по категориям
                if cat_col:
                    self.log("🔮 Прогнозируем продажи по категориям...")
                    cat_forecast_list = []
                    for cat, group in self.df.groupby(cat_col):
                        df_fc = forecast_group(group, cat, cat_col)
                        if df_fc is not None:
                            cat_forecast_list.append(df_fc)
                    if cat_forecast_list:
                        cat_forecast_full = pd.concat(cat_forecast_list, ignore_index=True)
                        self.results_storage['Прогноз_по_категориям_30дней'] = cat_forecast_full
                        self.log(f"   Готово: {len(cat_forecast_list)} категорий с прогнозом")

                # Прогноз по продуктам (только топ-20 по продажам, чтобы не перегружать)
                if lbl_col and not prod_stats.empty:
                    self.log("🔮 Прогнозируем продажи по продуктам (топ-20)...")
                    top_products = prod_stats.nlargest(20, 'Суммарные_Продажи')[lbl_col].tolist()
                    prod_forecast_list = []
                    for prod in top_products:
                        group = self.df[self.df[lbl_col] == prod]
                        df_fc = forecast_group(group, prod, lbl_col)
                        if df_fc is not None:
                            prod_forecast_list.append(df_fc)
                    if prod_forecast_list:
                        prod_forecast_full = pd.concat(prod_forecast_list, ignore_index=True)
                        self.results_storage['Прогноз_по_продуктам_Топ20_30дней'] = prod_forecast_full
                        self.log(f"   Готово: {len(prod_forecast_list)} продуктов с прогнозом")

            except Exception as e:
                self.log(f"⚠ Ошибка при прогнозировании: {e}")                

        self.btn_export.config(state="normal")
        self.log("✅ Анализ завершён! Все данные готовы для экспорта.")

        self.build_dashboard(
            prod_stats=prod_stats,
            cat_stats=cat_stats,
            cat2_stats=cat2_stats,
            has_profit=has_profit,
            total_sales=total_sales,
            date_col=date_col,
            cat_col=cat_col,
            cat2_col=cat2_col,
            lbl_col=lbl_col
        )
        
        self.build_side_dashboard()
        self.last_dashboard_key = None
        self.build_side_dashboard()


    def build_dashboard(self, prod_stats, cat_stats, cat2_stats, has_profit, total_sales, date_col, cat_col, cat2_col, lbl_col):
            # Удаляем старые виджеты (полная перерисовка каркаса)
            for widget in self.frame_dashboard.winfo_children():
                widget.destroy()
            self.canvas_dict.clear()
            
            # Сохраняем ссылки
            self.date_col = date_col
            self.cat_col = cat_col
            self.cat2_col = cat2_col
            self.cat2_listbox = None
            self.lbl_col = lbl_col
            self.df_filtered = self.df.copy()
    
            # ================= ФИЛЬТРЫ (Grid Layout) =================
            filter_frame = ttk.LabelFrame(self.frame_dashboard, text="Панель фильтров", padding=10)
            filter_frame.pack(fill="x", padx=10, pady=5)
            
            # --- Колонка 0: ДАТЫ ---
            col_date = ttk.Frame(filter_frame)
            col_date.grid(row=0, column=0, padx=10, sticky="nw")
            
            self.filter_date_from = tk.StringVar()
            self.filter_date_to = tk.StringVar()
            
            if date_col:
                ttk.Label(col_date, text="Период:", font=("Arial", 9, "bold")).pack(anchor="w", pady=(0, 5))
                
                # Дата С
                ttk.Label(col_date, text="C:").pack(anchor="w")
                if HAS_TKCALENDAR:
                    de_from = DateEntry(col_date, width=12, background='darkblue',
                                        foreground='white', borderwidth=2,
                                        textvariable=self.filter_date_from,
                                        date_pattern='yyyy-mm-dd')
                    de_from.delete(0, "end")
                    de_from.pack(anchor="w", pady=2)
                    # ВАЖНО: Автообновление при выборе даты
                    de_from.bind("<<DateEntrySelected>>", lambda e: self.apply_filters_and_update())
                else:
                    e_from = ttk.Entry(col_date, textvariable=self.filter_date_from, width=15)
                    e_from.pack(anchor="w", pady=2)
                    e_from.bind("<Return>", lambda e: self.apply_filters_and_update())
                    ttk.Label(col_date, text="(yyyy-mm-dd)", font=("Arial", 7)).pack(anchor="w")
    
                # Дата ПО
                ttk.Label(col_date, text="По:").pack(anchor="w")
                if HAS_TKCALENDAR:
                    de_to = DateEntry(col_date, width=12, background='darkblue',
                                      foreground='white', borderwidth=2,
                                      textvariable=self.filter_date_to,
                                      date_pattern='yyyy-mm-dd')
                    de_to.delete(0, "end")
                    de_to.pack(anchor="w", pady=2)
                    # ВАЖНО: Автообновление при выборе даты
                    de_to.bind("<<DateEntrySelected>>", lambda e: self.apply_filters_and_update())
                else:
                    e_to = ttk.Entry(col_date, textvariable=self.filter_date_to, width=15)
                    e_to.pack(anchor="w", pady=2)
                    e_to.bind("<Return>", lambda e: self.apply_filters_and_update())
            
            # Разделитель
            ttk.Separator(filter_frame, orient="vertical").grid(row=0, column=1, sticky="ns", padx=5)
    
            # --- Колонка 2: КАТЕГОРИЯ 1 ---
            col_cat1 = ttk.Frame(filter_frame)
            col_cat1.grid(row=0, column=2, padx=10, sticky="nw")
            
            self.category_listbox = None
            if cat_col:
                ttk.Label(col_cat1, text=f"{cat_col}:", font=("Arial", 9, "bold")).pack(anchor="w", pady=(0, 5))
                
                self.category_listbox = tk.Listbox(col_cat1, selectmode="extended", height=6, width=25, exportselection=False)
                scr1 = ttk.Scrollbar(col_cat1, orient="vertical", command=self.category_listbox.yview)
                self.category_listbox.configure(yscrollcommand=scr1.set)
                
                self.category_listbox.pack(side="left", fill="y")
                scr1.pack(side="right", fill="y")
                
                # Заполнение
                cats = sorted(self.df_filtered[cat_col].dropna().unique())
                for c in cats:
                    self.category_listbox.insert(tk.END, c)
                
                # Автообновление при клике
                self.category_listbox.bind("<<ListboxSelect>>", lambda *args: self.apply_filters_and_update())
    
            # --- Колонка 3: КАТЕГОРИЯ 2 ---
            col_cat2 = ttk.Frame(filter_frame)
            col_cat2.grid(row=0, column=3, padx=10, sticky="nw")
    
            self.cat2_listbox = None
            if cat2_col:
                ttk.Label(col_cat2, text=f"{cat2_col}:", font=("Arial", 9, "bold")).pack(anchor="w", pady=(0, 5))
                
                self.cat2_listbox = tk.Listbox(col_cat2, selectmode="extended", height=6, width=25, exportselection=False)
                scr2 = ttk.Scrollbar(col_cat2, orient="vertical", command=self.cat2_listbox.yview)
                self.cat2_listbox.configure(yscrollcommand=scr2.set)
                
                self.cat2_listbox.pack(side="left", fill="y")
                scr2.pack(side="right", fill="y")
                
                # Заполнение
                cats2 = sorted(self.df_filtered[cat2_col].dropna().unique())
                for c in cats2:
                    self.cat2_listbox.insert(tk.END, c)
                
                # Автообновление при клике
                self.cat2_listbox.bind("<<ListboxSelect>>", lambda *args: self.apply_filters_and_update())
    
            # --- Колонка 4: ПРОДУКТЫ ---
            col_prod = ttk.Frame(filter_frame)
            col_prod.grid(row=0, column=4, padx=10, sticky="nw")
            
            self.filter_product = tk.StringVar(value="Все")
            if lbl_col:
                ttk.Label(col_prod, text="Продукт:", font=("Arial", 9, "bold")).pack(anchor="w", pady=(0, 5))
                
                products = ["Все"] + sorted(self.df_filtered[lbl_col].dropna().unique().tolist())
                cb_prod = ttk.Combobox(col_prod, textvariable=self.filter_product, values=products, state="readonly", width=25)
                cb_prod.pack(anchor="w")
                
                # Автообновление при выборе
                self.filter_product.trace_add("write", lambda *args: self.apply_filters_and_update())
    
            # --- Колонка 5: КНОПКИ ---
            col_btn = ttk.Frame(filter_frame)
            col_btn.grid(row=0, column=5, padx=20, sticky="e")
            
            ttk.Button(col_btn, text="❌ Сброс", command=self.reset_filters, width=15).pack(pady=5)
    
            # ================= NOTEBOOK (Контейнер графиков) =================
            self.dashboard_nb = ttk.Notebook(self.frame_dashboard)
            self.dashboard_nb.pack(fill="both", expand=True, padx=10, pady=10)
            
            # Строим графики первый раз
            self.rebuild_dashboard(prod_stats, cat_stats, cat2_stats, has_profit, total_sales)
    
            self.log("Дашборд успешно построен! 📊")
        
    def kpi_card(self, parent, title, value, delta=None, positive=True):
            # card теперь растягивается во все стороны внутри своего родителя
            card = tk.Frame(parent, bg="white", highlightbackground=DASH_BORDER,
                            highlightthickness=1)
            card.pack(fill="both", expand=True, padx=3, pady=3) # Уменьшили padx/pady для компактности
        
            tk.Label(card, text=title, bg="white",
                     fg="#666", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(6, 0))
        
            tk.Label(card, text=value, bg="white",
                     fg="#000", font=("Arial", 14, "bold")).pack(anchor="w", padx=10) # Чуть меньше шрифт, чтобы влезло
        
            if delta is not None:
                color = DASH_GREEN if positive else DASH_RED
                sign = "▲" if positive else "▼"
                tk.Label(
                    card,
                    text=f"{sign} {delta}",
                    bg="white",
                    fg=color,
                    font=("Arial", 9, "bold")
                ).pack(anchor="w", padx=10, pady=(0, 6))
            
    def barh_top_bottom(self, parent, series, title):
        """
        series: pandas Series (index = название, values = сумма)
        """
        # ДИНАМИЧЕСКИЙ РАЗМЕР В ЗАВИСИМОСТИ ОТ ТИПА ГРАФИКА
        if "продукт" in title.lower():
            fig, ax = plt.subplots(figsize=(3.5, 1.8))  # Шире для продуктов
        else:
            fig, ax = plt.subplots(figsize=(3.0, 1.8))  # Стандартно для категорий
    
        top = series.sort_values(ascending=False).head(5)
        bottom = series.sort_values().head(5)
    
        labels = list(top.index) + list(bottom.index)
        values = list(top.values) + list(bottom.values)
    
        colors = [DASH_ORANGE]*len(top) + ["#cccccc"]*len(bottom)
    
        y_pos = np.arange(len(labels))
    
        ax.barh(y_pos, values, color=colors)
        ax.set_yticks(y_pos)
        
        # Устанавливаем шрифты как в графике продаж за 12 месяцев
        ax.set_yticklabels(labels, fontsize=6)  # Еще меньше размер для компактности
        ax.invert_yaxis()
    
        ax.set_title(title, fontsize=8, pad=8)  # Уменьшаем шрифт заголовка
        ax.grid(axis="x", alpha=0.3)
    
        # Подписи значений - устанавливаем те же настройки шрифтов
        for i, v in enumerate(values):
            ax.text(v, i, f"{v:,.0f}".replace(',', ' '), 
                    va="center", 
                    ha="left", 
                    fontsize=6,  # Уменьшаем размер шрифта
                    fontweight='bold')
    
        # Настройка шрифтов делений осей
        ax.tick_params(axis='both', which='major', labelsize=6)
        
        fig.tight_layout()
    
        canvas = FigureCanvasTkAgg(fig, parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        
    def build_side_dashboard(self):
            if self.df is None or self.date_col is None:
                return
        
            # ===== КЭШ-КЛЮЧ =====
            key = (
                len(self.df),
                self.df[self.date_col].max(),
                self.df["Продажи"].sum()
            )
        
            if key == self.last_dashboard_key:
                return
        
            self.last_dashboard_key = key
        
            for w in self.dashboard_side.winfo_children():
                w.destroy()
        
            df = self.df.copy()
            df[self.date_col] = pd.to_datetime(df[self.date_col], errors="coerce")
            df = df.dropna(subset=[self.date_col])
        
            last_date = df[self.date_col].max()
        
            df_12m = df[df[self.date_col] >= last_date - pd.DateOffset(months=12)]
            df_prev = df[
                (df[self.date_col] < last_date - pd.DateOffset(months=12)) &
                (df[self.date_col] >= last_date - pd.DateOffset(months=24))
            ]
        
            rev_now = df_12m["Продажи"].sum()
            rev_prev = df_prev["Продажи"].sum()
            delta_rev = ((rev_now - rev_prev) / rev_prev * 100) if rev_prev else 0
        
            prof_now = df_12m["Прибыль"].sum() if "Прибыль" in df.columns else None
            prof_prev = df_prev["Прибыль"].sum() if "Прибыль" in df.columns else None
            delta_prof = ((prof_now - prof_prev) / prof_prev * 100) if (prof_prev and prof_prev != 0) else 0
        
            # ===== KPI (В ОДНУ СТРОКУ) =====
            kpi_container = tk.Frame(self.dashboard_side)
            kpi_container.pack(fill="x", padx=3, pady=3)
    
            # Левая колонка для Выручки
            kpi_left = tk.Frame(kpi_container)
            kpi_left.pack(side="left", fill="both", expand=True)
    
            self.kpi_card(
                kpi_left,
                "Выручка (12 мес)",
                "{:,.0f}".format(rev_now).replace(',', ' '),
                "{:+.1f}%".format(delta_rev).replace(',', ' '),
                delta_rev >= 0
            )
        
            # Правая колонка для Прибыли (если есть)
            if prof_now is not None:
                kpi_right = tk.Frame(kpi_container)
                kpi_right.pack(side="left", fill="both", expand=True) # Тоже side=left
    
                # --- ИСПРАВЛЕНИЕ ОШИБКИ ---
                # Раньше здесь использовались rev_now и delta_rev. 
                # Заменили на prof_now и delta_prof
                self.kpi_card(
                    kpi_right,
                    "Прибыль (12 мес)",
                    "{:,.0f}".format(prof_now).replace(',', ' '),  # <-- Исправлено
                    "{:+.1f}%".format(delta_prof).replace(',', ' '), # <-- Исправлено
                    delta_prof >= 0
                )
        
            # ===== ГРАФИК ВЫРУЧКИ ПО МЕСЯЦАМ =====
            fig, ax = plt.subplots(figsize=(5.2, 2.2))
            
            # Группируем данные
            ts = df_12m.groupby(pd.Grouper(key=self.date_col, freq="MS"))["Продажи"].sum()
            
            if not ts.empty:
                ts.index = pd.to_datetime(ts.index)
                
                ax.plot(ts.index, ts.values, marker="o", color=DASH_ORANGE, linewidth=2)
                ax.fill_between(ts.index, ts.values, alpha=0.15, color=DASH_ORANGE)
            
                if len(ts) > 1:
                    ax.set_xlim(left=ts.index[1])
                
                y_offset = ts.values.max() * 0.15 if not ts.empty else 0
                
                for x, y in zip(ts.index, ts.values):
                    if len(ts) > 1 and x == ts.index[0]:
                        continue
                        
                    ax.text(x, y + y_offset, f'{y:,.0f}'.replace(',', ' '), 
                            fontsize=7, 
                            ha='left', 
                            va='bottom',
                            fontweight='bold')
            
                ax.set_title("Выручка по месяцам", fontsize=9, pad=10)
                ax.grid(True, linestyle='--', alpha=0.6)
                
                if not ts.empty:
                    ax.set_ylim(bottom=0, top=ts.values.max() * 1.3)
            
                ax.tick_params(axis='both', which='major', labelsize=7)
            
            fig.tight_layout()
            canvas = FigureCanvasTkAgg(fig, self.dashboard_side)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="x", padx=6, pady=6)
            
            # ===== СЕТКА ДЛЯ ТОП/БОТТОМ =====
            graph_container = tk.Frame(self.dashboard_side)
            graph_container.pack(fill="x", padx=6, pady=6)
            
            left_col = tk.Frame(graph_container)
            left_col.pack(side="left", fill="both", expand=True, padx=(0, 3))
            
            right_col = tk.Frame(graph_container)
            right_col.pack(side="right", fill="both", expand=True, padx=(3, 0))
            
            # ===== TOP / BOTTOM ПРОДУКТЫ =====
            if self.lbl_col:
                prod_series = (
                    df_12m
                    .groupby(self.lbl_col)["Продажи"]
                    .sum()
                    .sort_values()
                )
                self.barh_top_bottom(left_col, prod_series, "Top / Bottom 5 продуктов")
            else:
                tk.Label(left_col, text="Продукты не выбраны", 
                         font=("Arial", 9), fg="gray").pack(fill="both", expand=True)
            
            # ===== TOP / BOTTOM КАТЕГОРИИ =====
            if self.cat_col:
                cat_series = (
                    df_12m
                    .groupby(self.cat_col)["Продажи"]
                    .sum()
                    .sort_values()
                )
                self.barh_top_bottom(right_col, cat_series, "Top / Bottom 5 категорий")
            else:
                tk.Label(right_col, text="Категории не выбраны", 
                         font=("Arial", 9), fg="gray").pack(fill="both", expand=True)

    def export_excel(self):
        if not self.results_storage:
            messagebox.showwarning("Нет данных", "Сначала выполните анализ.")
            return
    
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файл", "*.xlsx"), ("CSV файл", "*.csv")],
            title="Сохранить расширенный отчёт"
        )
        if not save_path:
            return
    
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.formatting.rule import CellIsRule
            
            # Создаем новый Workbook
            wb = Workbook()
            
            # Удаляем дефолтный лист, если он есть
            if 'Sheet' in wb.sheetnames:
                ws_default = wb['Sheet']
                wb.remove(ws_default)
            
            # ====== 1. СОЗДАЕМ ИНТЕРАКТИВНОЕ СОДЕРЖАНИЕ ======
            ws_toc = wb.create_sheet(title="📑 Содержание", index=0)
            
            # Заголовок содержания
            ws_toc['A1'] = "📊 СОДЕРЖАНИЕ ОТЧЁТА"
            ws_toc['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws_toc['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws_toc.merge_cells('A1:D1')
            ws_toc['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Заголовки колонок
            headers = ["№", "Название листа", "Описание", "Быстрый переход"]
            for col, header in enumerate(headers, 1):
                cell = ws_toc.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Собираем все доступные листы
            sheet_order = [
                'Общие_показатели',
                'Метрики_по_категориям',
                'Метрики_по_категории_2',
                'Метрики_по_продуктам',
                'TOP_5_Категории_Продажи',
                'TOP_5_Продукты_Продажи',
                'BOTTOM_5_Продукты_Продажи',
                'TOP_5_Категории_Прибыль',
                'TOP_5_Продукты_Прибыль',
                'ABC_XYZ_Продукты_Детально',
                'ABC_XYZ_Продукты_Статистика',
                'ABC_XYZ_Категории',
                'Динамика_по_месяцам',
                'Динамика_по_годам',
                'Сезонность_по_месяцам',
                'Сезонность_детально',
                'Сезонность_по_годам_и_месяцам',
                'Эластичность_спроса_детально',
                'Эластичность_спроса_сводка',
                'Корреляция_цена_количество',
                'Прогноз_по_категориям_30дней',
                'Прогноз_по_продуктам_Топ20_30дней'
            ]
            
            # Создаем список всех листов
            all_sheets_info = []
            
            # 1. Листы из sheet_order
            for sheet_name in sheet_order:
                if sheet_name in self.results_storage:
                    all_sheets_info.append(sheet_name)
            
            # 2. Остальные листы
            for sheet_name in self.results_storage:
                if sheet_name not in sheet_order and sheet_name not in ['Содержание', '📑 Содержание']:
                    all_sheets_info.append(sheet_name)
            
            # Заполняем таблицу содержания
            row = 4
            sheet_name_map = {}  # Для отслеживания имен листов
            
            for i, sheet_name in enumerate(all_sheets_info, 1):
                df = self.results_storage[sheet_name]
                description = self._get_single_sheet_description(sheet_name)
                
                # Создаем безопасное имя для листа
                safe_name = sheet_name[:27]  # Оставляем место для возможного суффикса
                
                # Убедимся, что имя уникально
                base_name = safe_name
                counter = 1
                original_safe_name = safe_name
                while safe_name in wb.sheetnames:
                    safe_name = f"{base_name}_{counter}"
                    counter += 1
                
                sheet_name_map[sheet_name] = safe_name
                
                # Записываем данные в строку содержания
                # 1. Номер
                ws_toc.cell(row=row, column=1, value=i)
                ws_toc.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                
                # 2. Название листа (с гиперссылкой)
                cell_name = ws_toc.cell(row=row, column=2, value=safe_name)
                cell_name.font = Font(color="0563C1", underline="single")
                cell_name.hyperlink = f"#'{safe_name}'!A1"
                
                # 3. Описание
                ws_toc.cell(row=row, column=3, value=description)
                
                # 4. Кнопка перехода (текстовая)
                cell_button = ws_toc.cell(row=row, column=4, value="▶ Перейти")
                cell_button.font = Font(color="FFFFFF", bold=True)
                cell_button.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                cell_button.alignment = Alignment(horizontal='center')
                cell_button.hyperlink = f"#'{safe_name}'!A1"
                
                # Добавляем границы для всей строки
                for col in range(1, 5):
                    ws_toc.cell(row=row, column=col).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Чередование цветов строк
                if row % 2 == 0:
                    for col in range(1, 5):
                        ws_toc.cell(row=row, column=col).fill = PatternFill(
                            start_color="F2F2F2", 
                            end_color="F2F2F2", 
                            fill_type="solid"
                        )
                
                row += 1
            
            # Добавляем строку с итогом
            ws_toc.cell(row=row, column=1, value="ИТОГО:")
            ws_toc.cell(row=row, column=1).font = Font(bold=True)
            ws_toc.cell(row=row, column=2, value=f"{len(all_sheets_info)} листов")
            ws_toc.cell(row=row, column=2).font = Font(bold=True)
            ws_toc.merge_cells(f'A{row}:B{row}')
            
            # Добавляем кнопку "Вернуться к содержанию" на каждом листе
            return_button_text = "🔙 К содержанию"
            return_button_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            # ====== 2. СОЗДАЕМ ЛИСТЫ С ДАННЫМИ ======
            for sheet_name, safe_name in sheet_name_map.items():
                df = self.results_storage[sheet_name]
                
                # Создаем лист
                ws = wb.create_sheet(title=safe_name)
                
                # Добавляем кнопку возврата в первую строку
                cell_return = ws['A1']
                cell_return.value = return_button_text
                cell_return.font = Font(color="FFFFFF", bold=True)
                cell_return.fill = return_button_fill
                cell_return.alignment = Alignment(horizontal='center')
                cell_return.hyperlink = f"#'📑 Содержание'!A1"
                
                # Записываем заголовок листа
                ws['A2'] = sheet_name
                ws['A2'].font = Font(size=12, bold=True, color="1F497D")
                
                # Записываем данные (начиная с 4 строки, чтобы оставить место для заголовков)
                for r_idx, row_data in enumerate(df.itertuples(index=False), start=4):
                    for c_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Форматирование для числовых значений
                        if isinstance(value, (int, float)):
                            if abs(value) >= 1000:
                                cell.number_format = '#,##0'
                            else:
                                cell.number_format = '0.00'
                        
                        # Добавляем границы
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                
                # Записываем заголовки колонок
                for c_idx, col_name in enumerate(df.columns, start=1):
                    cell = ws.cell(row=3, column=c_idx, value=col_name)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Автонастройка ширины колонок
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Закрепляем первые 3 строки (заголовки)
                ws.freeze_panes = ws['A4']
            
            # Настраиваем ширину колонок в содержании
            ws_toc.column_dimensions['A'].width = 6      # №
            ws_toc.column_dimensions['B'].width = 28     # Название листа
            ws_toc.column_dimensions['C'].width = 45     # Описание
            ws_toc.column_dimensions['D'].width = 12     # Быстрый переход
            
            # Добавляем автофильтр к заголовкам содержания
            ws_toc.auto_filter.ref = f"A3:D{row-1}"
            
            # Закрепляем заголовки содержания
            ws_toc.freeze_panes = ws_toc['A4']
            
            # ====== 3. СОХРАНЯЕМ ФАЙЛ ======
            wb.save(save_path)
            
            self.log(f"💾 Расширенный отчёт сохранён: {os.path.basename(save_path)}")
            messagebox.showinfo(
                "Успешный экспорт", 
                f"✅ Отчёт успешно сохранён!\n\n"
                f"📄 Файл: {os.path.basename(save_path)}\n"
                f"📊 Листов: {len(all_sheets_info) + 1}\n"
            )
            
            # Предложение открыть файл
            if messagebox.askyesno("Открыть файл", "Хотите открыть сохранённый файл для проверки?"):
                if os.name == 'nt':  # Windows
                    os.startfile(save_path)
                elif os.name == 'posix':  # macOS, Linux
                    import subprocess
                    subprocess.call(('open', save_path))
                
        except ImportError:
            # Fallback на стандартный метод, если openpyxl не установлен
            messagebox.showwarning(
                "Установите openpyxl", 
                "Для интерактивного содержания требуется библиотека openpyxl.\n"
                "Установите: pip install openpyxl\n\n"
                "Будет создан стандартный отчёт без гиперссылок."
            )
            self._export_excel_fallback(save_path)
            
        except PermissionError:
            messagebox.showerror(
                "Ошибка доступа", 
                f"❌ Нет прав для записи файла.\n"
                f"Возможно, файл открыт в другой программе."
            )
        except Exception as e:
            messagebox.showerror(
                "Ошибка экспорта", 
                f"❌ Не удалось сохранить файл:\n\n{str(e)}"
            )
    
    def _export_excel_fallback(self, save_path):
        """Резервный метод экспорта без интерактивных функций"""
        try:
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                # Создаем содержание
                sheet_order = [
                    'Общие_показатели',
                    'Метрики_по_категориям',
                    'Метрики_по_категории_2',
                    'Метрики_по_продуктам',
                    'TOP_5_Категории_Продажи',
                    'TOP_5_Продукты_Продажи',
                    'BOTTOM_5_Продукты_Продажи',
                    'TOP_5_Категории_Прибыль',
                    'TOP_5_Продукты_Прибыль',
                    'ABC_XYZ_Продукты_Детально',
                    'ABC_XYZ_Продукты_Статистика',
                    'ABC_XYZ_Категории',
                    'Динамика_по_месяцам',
                    'Динамика_по_годам',
                    'Сезонность_по_месяцам',
                    'Сезонность_детально',
                    'Сезонность_по_годам_и_месяцам',
                    'Эластичность_спроса_детально',
                    'Эластичность_спроса_сводка',
                    'Корреляция_цена_количество',
                    'Прогноз_по_категориям_30дней',
                    'Прогноз_по_продуктам_Топ20_30дней'
                ]
                
                all_sheets_info = []
                for sheet_name in sheet_order:
                    if sheet_name in self.results_storage:
                        all_sheets_info.append(sheet_name)
                
                for sheet_name in self.results_storage:
                    if sheet_name not in sheet_order:
                        all_sheets_info.append(sheet_name)
                
                # Создаем оглавление
                toc_data = []
                for i, sheet_name in enumerate(all_sheets_info, 1):
                    description = self._get_single_sheet_description(sheet_name)
                    toc_data.append({
                        'Номер': i,
                        'Название листа': sheet_name[:31],
                        'Описание': description
                    })
                
                toc_df = pd.DataFrame(toc_data)
                toc_df.to_excel(writer, sheet_name='Содержание', index=False)
                
                # Экспортируем остальные листы
                for sheet_name in all_sheets_info:
                    df = self.results_storage[sheet_name]
                    safe_name = sheet_name[:31]
                    df.to_excel(writer, sheet_name=safe_name, index=False)
            
            messagebox.showinfo(
                "Экспорт завершён", 
                f"Создан отчёт без интерактивных функций.\n"
                f"Установите openpyxl для улучшенного экспорта."
            )
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать файл: {e}")
    
    def _get_single_sheet_description(self, sheet_name):
        """Возвращает описание для конкретного листа"""
        descriptions = {
            'Общие_показатели': 'Общая статистика по всем данным',
            'Метрики_по_категориям': 'Детальная статистика по категориям',
            'Метрики_по_категории_2': 'Статистика по дополнительной категории',
            'Метрики_по_продуктам': 'Детальная статистика по продуктам',
            'TOP_5_Категории_Продажи': 'Топ-5 категорий по продажам',
            'TOP_5_Продукты_Продажи': 'Топ-5 продуктов по продажам',
            'BOTTOM_5_Продукты_Продажи': 'Худшие 5 продуктов по продажам',
            'TOP_5_Категории_Прибыль': 'Топ-5 категорий по прибыли',
            'TOP_5_Продукты_Прибыль': 'Топ-5 продуктов по прибыли',
            'ABC_XYZ_Продукты_Детально': 'ABC-XYZ анализ продуктов (детально)',
            'ABC_XYZ_Продукты_Статистика': 'ABC-XYZ анализ продуктов (статистика)',
            'ABC_XYZ_Категории': 'ABC-XYZ анализ по категориям',
            'Динамика_по_месяцам': 'Динамика продаж по месяцам',
            'Динамика_по_годам': 'Динамика продаж по годам',
            'Сезонность_по_месяцам': 'Сезонность продаж по месяцам',
            'Сезонность_детально': 'Детальный анализ сезонности',
            'Сезонность_по_годам_и_месяцам': 'Сезонность по годам и месяцам',
            'Эластичность_спроса_детально': 'Эластичность спроса по товарам/категориям',
            'Эластичность_спроса_сводка': 'Сводная статистика по эластичности',
            'Корреляция_цена_количество': 'Корреляция между ценой и количеством',
            'Прогноз_по_категориям_30дней': 'Прогноз продаж по категориям на 30 дней',
            'Прогноз_по_продуктам_Топ20_30дней': 'Прогноз продаж по топ-20 продуктам',
        }
        
        for key, desc in descriptions.items():
            if key == sheet_name or key in sheet_name:
                return desc
        
        return 'Дополнительные данные'

   
    def apply_filters_and_update(self):
            """Применяет фильтры и обновляет ТОЛЬКО графики"""
            try:
                # Сбрасываем к исходным данным
                self.df = self.df_filtered.copy()
                
                # 1. Фильтрация по датам
                if self.date_col and hasattr(self, 'filter_date_from') and hasattr(self, 'filter_date_to'):
                    self.df[self.date_col] = pd.to_datetime(self.df[self.date_col], errors='coerce')
                    
                    val_from = self.filter_date_from.get()
                    val_to = self.filter_date_to.get()
    
                    if val_from:
                        try:
                            from_date = pd.to_datetime(val_from)
                            self.df = self.df[self.df[self.date_col] >= from_date]
                        except: pass
                    
                    if val_to:
                        try:
                            to_date = pd.to_datetime(val_to)
                            self.df = self.df[self.df[self.date_col] <= to_date]
                        except: pass
                
                # 2. Фильтрация по категориям
                if self.cat_col and self.category_listbox:
                    # curselection возвращает индексы, берем текст
                    selected_indices = self.category_listbox.curselection()
                    selected_cats = [self.category_listbox.get(i) for i in selected_indices]
                    if selected_cats:
                        self.df = self.df[self.df[self.cat_col].isin(selected_cats)]
                        
                # 3. Фильтрация по Категории 2
                if self.cat2_col and hasattr(self, 'cat2_listbox') and self.cat2_listbox:
                    selected_indices2 = self.cat2_listbox.curselection()
                    selected_cats2 = [self.cat2_listbox.get(i) for i in selected_indices2]
                    if selected_cats2:
                        self.df = self.df[self.df[self.cat2_col].isin(selected_cats2)]
                
                # 4. Фильтрация по продукту
                if self.lbl_col and hasattr(self, 'filter_product'):
                    val_prod = self.filter_product.get()
                    if val_prod and val_prod != "Все":
                        self.df = self.df[self.df[self.lbl_col] == val_prod]
                
                # --- ПЕРЕСЧЕТ СТАТИСТИКИ ---
                total_sales = self.df['Продажи'].sum() if 'Продажи' in self.df.columns else 0
                has_profit = 'Прибыль' in self.df.columns
                
                prod_stats = pd.DataFrame()
                cat_stats = pd.DataFrame()
                cat2_stats = pd.DataFrame()
                
                # Статистика по продуктам
                if self.lbl_col:
                    prod_stats = self.df.groupby(self.lbl_col).agg(
                        Суммарные_Продажи=('Продажи', 'sum'),
                        Средние_Продажи=('Продажи', 'mean'),
                        Количество_транзакций=('Продажи', 'count')
                    ).reset_index()
                
                # Статистика по категории 1
                if self.cat_col:
                    cat_stats = self.df.groupby(self.cat_col).agg(
                        Суммарные_Продажи=('Продажи', 'sum'),
                        Средние_Продажи=('Продажи', 'mean'),
                        Количество_транзакций=('Продажи', 'count')
                    ).reset_index()
                
                # Статистика по категории 2
                if self.cat2_col:
                    cat2_stats = self.df.groupby(self.cat2_col).agg(
                        Суммарные_Продажи=('Продажи', 'sum'),
                        Средние_Продажи=('Продажи', 'mean'),
                        Количество_транзакций=('Продажи', 'count')
                    ).reset_index()
                
                # --- ОБНОВЛЕНИЕ ГРАФИКОВ ---
                # Вызываем rebuild_dashboard, который перерисует только вкладки Notebook,
                # но НЕ тронет фильтры
                self.rebuild_dashboard(prod_stats, cat_stats, cat2_stats, has_profit, total_sales)
                
            except Exception as e:
                print(f"Ошибка при фильтрации: {e}")
            
            self.build_side_dashboard()
            self.last_dashboard_key = None
            self.build_side_dashboard()

    
    def rebuild_dashboard(self, prod_stats, cat_stats, cat2_stats, has_profit, total_sales):
            """Перестраивает графики дашборда с новыми данными в едином стиле"""
            # Удаляем старые вкладки
            for tab in self.dashboard_nb.tabs():
                self.dashboard_nb.forget(tab)
            
            dashboard_nb = self.dashboard_nb
            
            # Общая функция стиля для осей (убираем лишние рамки, делаем сетку)
            def style_ax(ax, grid_axis='y'):
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.spines['left'].set_color('#888888')
                ax.spines['bottom'].set_color('#888888')
                ax.tick_params(colors='#333333', labelsize=8)
                ax.grid(True, axis=grid_axis, linestyle='--', alpha=0.6, zorder=0)
                ax.set_facecolor('white')
    
            # ====== ВКЛАДКА 1: ОБЗОР ======
            tab_overview = ttk.Frame(dashboard_nb)
            dashboard_nb.add(tab_overview, text="Обзор")
            
            fig1 = plt.Figure(figsize=(12, 6), facecolor='white')
            ax1 = fig1.add_subplot(121)
            ax2 = fig1.add_subplot(122)
            
            # --- Круговая диаграмма (Pie) ---
            if self.cat_col and not cat_stats.empty:
                pie_data = cat_stats[cat_stats['Суммарные_Продажи'] > 0].copy()
                if not pie_data.empty:
                    # Используем пастельную палитру, так как оранжевый тут сольется
                    colors = plt.cm.Pastel1(np.linspace(0, 1, len(pie_data)))
                    wedges, texts, autotexts = ax1.pie(
                        pie_data['Суммарные_Продажи'], 
                        labels=pie_data[self.cat_col], 
                        autopct='%1.1f%%', 
                        startangle=90,
                        colors=colors,
                        wedgeprops={'edgecolor': 'white'}
                    )
                    plt.setp(texts, size=8)
                    plt.setp(autotexts, size=8, weight="bold", color="white")
                    ax1.set_title("Доли категорий в выручке", fontsize=10, fontweight='bold', pad=10)
                else:
                    ax1.text(0.5, 0.5, "Нет данных > 0", ha='center')
                    ax1.axis('off')
            else:
                ax1.text(0.5, 0.5, "Категории не выбраны", ha='center')
                ax1.axis('off')
            
            # --- Treemap ---
            if self.cat_col and not cat_stats.empty and HAS_SQUARIFY:
                valid_cat = cat_stats[cat_stats['Суммарные_Продажи'] > 0].copy()
                if len(valid_cat) > 0:
                    sizes = valid_cat['Суммарные_Продажи'].tolist()
                    labels = [f"{n}\n{v:,.0f}".replace(',', ' ') for n, v in zip(valid_cat[self.cat_col], sizes)]
                    colors = plt.cm.Set3(np.linspace(0, 1, len(valid_cat)))
                    
                    try:
                        squarify.plot(sizes=sizes, label=labels, color=colors, alpha=0.7, ax=ax2, 
                                      text_kwargs={'fontsize': 8, 'color': '#333333'}, bar_kwargs={'linewidth': 1, 'edgecolor': 'white'})
                        ax2.axis('off')
                        ax2.set_title("Карта категорий (Treemap)", fontsize=10, fontweight='bold', pad=10)
                    except Exception as e:
                        ax2.text(0.5, 0.5, "Ошибка построения", ha='center')
                        ax2.axis('off')
                else:
                    ax2.axis('off')
            else:
                ax2.text(0.5, 0.5, "Treemap недоступен", ha='center')
                ax2.axis('off')
            
            canvas1 = FigureCanvasTkAgg(fig1, tab_overview)
            canvas1.draw()
            canvas1.get_tk_widget().pack(fill="both", expand=True)
            
            # ====== ВКЛАДКА 2: КАТЕГОРИИ (Барчарт) ======
            if self.cat_col and not cat_stats.empty:
                tab_cats = ttk.Frame(dashboard_nb)
                dashboard_nb.add(tab_cats, text="Категории")
                
                fig3 = plt.Figure(figsize=(10, 6), facecolor='white')
                ax = fig3.add_subplot(111)
                cat_sorted = cat_stats.sort_values('Суммарные_Продажи', ascending=False)
                
                # Стиль: Оранжевые столбцы
                bars = ax.bar(range(len(cat_sorted)), cat_sorted['Суммарные_Продажи'], 
                              color=DASH_ORANGE, zorder=3, width=0.6)
                
                style_ax(ax, grid_axis='y') # Применяем стиль
                
                ax.set_title("Продажи по категориям", fontsize=10, fontweight='bold', pad=15)
                ax.set_xticks(range(len(cat_sorted)))
                ax.set_xticklabels(cat_sorted[self.cat_col], rotation=45, ha='right', fontsize=9)
                
                # Подписи значений
                y_offset = cat_sorted['Суммарные_Продажи'].max() * 0.02
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height + y_offset,
                           "{:,.0f}".format(height).replace(',', ' '), 
                           ha='center', va='bottom', fontsize=8, fontweight='bold', color='#333333')
                
                # Немного увеличим верхнюю границу, чтобы влезли подписи
                ax.set_ylim(top=cat_sorted['Суммарные_Продажи'].max() * 1.15)
                
                canvas3 = FigureCanvasTkAgg(fig3, tab_cats)
                canvas3.draw()
                canvas3.get_tk_widget().pack(fill="both", expand=True)
                
            # ====== ВКЛАДКА: КАТЕГОРИЯ 2 (Барчарт) ======
            if self.cat2_col and not cat2_stats.empty:
                tab_cat2 = ttk.Frame(dashboard_nb)
                dashboard_nb.add(tab_cat2, text=f"Анализ: {self.cat2_col}")
                
                fig_c2 = plt.Figure(figsize=(10, 6), facecolor='white')
                ax = fig_c2.add_subplot(111)
                
                c2_sorted = cat2_stats.sort_values('Суммарные_Продажи', ascending=False).head(15) 
                
                # Стиль: Оранжевые столбцы
                bars = ax.bar(range(len(c2_sorted)), c2_sorted['Суммарные_Продажи'], 
                              color=DASH_ORANGE, zorder=3, width=0.6)
                
                style_ax(ax, grid_axis='y') # Применяем стиль
                
                ax.set_title(f"Продажи по: {self.cat2_col}", fontsize=10, fontweight='bold', pad=15)
                ax.set_xticks(range(len(c2_sorted)))
                ax.set_xticklabels(c2_sorted[self.cat2_col], rotation=45, ha='right', fontsize=9)
                
                y_offset = c2_sorted['Суммарные_Продажи'].max() * 0.02
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height + y_offset,
                           "{:,.0f}".format(height).replace(',', ' '), 
                           ha='center', va='bottom', fontsize=8, fontweight='bold', color='#333333')
                
                ax.set_ylim(top=c2_sorted['Суммарные_Продажи'].max() * 1.15)
                
                canvas_c2 = FigureCanvasTkAgg(fig_c2, tab_cat2)
                canvas_c2.draw()
                canvas_c2.get_tk_widget().pack(fill="both", expand=True)
            
            # ====== ВКЛАДКА 3: ПРОДУКТЫ (Горизонтальный бар) ======
            if self.lbl_col and not prod_stats.empty:
                tab_prods = ttk.Frame(dashboard_nb)
                dashboard_nb.add(tab_prods, text="Продукты")
                
                top_prods = prod_stats.nlargest(10, 'Суммарные_Продажи')
                
                fig4 = plt.Figure(figsize=(12, 6), facecolor='white')
                ax = fig4.add_subplot(111)
                
                # Стиль: Оранжевые горизонтальные полосы
                bars = ax.barh(range(len(top_prods)), top_prods['Суммарные_Продажи'], 
                               color=DASH_ORANGE, zorder=3, height=0.6)
                
                style_ax(ax, grid_axis='x') # Сетка по X для горизонтального графика
                
                ax.set_title("Топ-10 продуктов по продажам", fontsize=10, fontweight='bold', pad=10)
                ax.set_yticks(range(len(top_prods)))
                ax.set_yticklabels(top_prods[self.lbl_col], fontsize=9)
                ax.invert_yaxis() # Сверху вниз
                
                # Подписи справа от баров
                x_offset = top_prods['Суммарные_Продажи'].max() * 0.01
                for bar in bars:
                    width = bar.get_width()
                    ax.text(width + x_offset, bar.get_y() + bar.get_height()/2.,
                           "{:,.0f}".format(width).replace(',', ' '), 
                           ha='left', va='center', fontsize=8, fontweight='bold', color='#333333')
                
                ax.set_xlim(right=top_prods['Суммарные_Продажи'].max() * 1.15)
                
                canvas4 = FigureCanvasTkAgg(fig4, tab_prods)
                canvas4.draw()
                canvas4.get_tk_widget().pack(fill="both", expand=True)
            
            # ====== ВКЛАДКА 4: ДИНАМИКА (Линейный график) ======
            if self.date_col and hasattr(self, 'df') and 'Продажи' in self.df.columns:
                try:
                    if pd.api.types.is_datetime64_any_dtype(self.df[self.date_col]):
                        ts = self.df.groupby(pd.Grouper(key=self.date_col, freq='M'))['Продажи'].sum().reset_index()
                        
                        if not ts.empty:
                            tab_time = ttk.Frame(dashboard_nb)
                            dashboard_nb.add(tab_time, text="Динамика")
                            
                            fig5 = plt.Figure(figsize=(12, 6), facecolor='white')
                            ax = fig5.add_subplot(111)
                            
                            # Основная линия - Оранжевая с заливкой (как в side_dashboard)
                            ax.plot(ts[self.date_col], ts['Продажи'], 
                                    marker='o', markersize=5, 
                                    color=DASH_ORANGE, linewidth=2, 
                                    label='Продажи', zorder=3)
                            
                            # Заливка под графиком
                            ax.fill_between(ts[self.date_col], ts['Продажи'], 
                                            color=DASH_ORANGE, alpha=0.15, zorder=2)
                            
                            # Скользящая средняя - Темно-серая пунктирная
                            if len(ts) > 3:
                                ax.plot(ts[self.date_col], ts['Продажи'].rolling(3, min_periods=1).mean(), 
                                        label='Тренд (3 мес)', linestyle='--', color='#555555', linewidth=1.5, zorder=4)
                            
                            style_ax(ax, grid_axis='both') # Сетка по X и Y
                            
                            ax.set_title("Динамика продаж по месяцам", fontsize=10, fontweight='bold', pad=10)
                            ax.legend(frameon=False, fontsize=9)
                            ax.tick_params(axis='x', rotation=45)
                            ax.set_ylim(bottom=0) # Начинаем с 0
                            
                            canvas5 = FigureCanvasTkAgg(fig5, tab_time)
                            canvas5.draw()
                            canvas5.get_tk_widget().pack(fill="both", expand=True)
                except Exception as e:
                    print(f"Ошибка при построении динамики: {e}")
            
            dashboard_nb.update_idletasks()
    
    def reset_filters(self):
            """Сброс всех фильтров"""
            if hasattr(self, 'filter_date_from'):
                self.filter_date_from.set("")
            if hasattr(self, 'filter_date_to'):
                self.filter_date_to.set("")
            if hasattr(self, 'filter_product'):
                self.filter_product.set("Все")
            if hasattr(self, 'category_listbox'):
                self.category_listbox.selection_clear(0, tk.END)
            if hasattr(self, 'cat2_listbox') and self.cat2_listbox:
                self.cat2_listbox.selection_clear(0, tk.END)
            
            # Возвращаемся к исходным данным
            self.df = self.df_filtered.copy()
            
            # Пересчитываем метрики
            total_sales = self.df['Продажи'].sum() if 'Продажи' in self.df.columns else 0
            has_profit = 'Прибыль' in self.df.columns
            
            prod_stats = pd.DataFrame()
            cat_stats = pd.DataFrame()
            cat2_stats = pd.DataFrame()  # --- ИСПРАВЛЕНИЕ: Добавили переменную
            
            if self.lbl_col:
                prod_stats = self.df.groupby(self.lbl_col).agg(
                    Суммарные_Продажи=('Продажи', 'sum'),
                    Средние_Продажи=('Продажи', 'mean'),
                    Медиана=('Продажи', 'median'),
                    StdDev=('Продажи', 'std'),
                    Количество_транзакций=('Продажи', 'count')
                ).reset_index()
            
            if self.cat_col:
                cat_stats = self.df.groupby(self.cat_col).agg(
                    Суммарные_Продажи=('Продажи', 'sum'),
                    Средние_Продажи=('Продажи', 'mean'),
                    Количество_транзакций=('Продажи', 'count')
                ).reset_index()
                
            # --- ИСПРАВЛЕНИЕ: Добавляем расчет для cat2_stats
            if self.cat2_col:
                cat2_stats = self.df.groupby(self.cat2_col).agg(
                    Суммарные_Продажи=('Продажи', 'sum'),
                    Средние_Продажи=('Продажи', 'mean'),
                    Количество_транзакций=('Продажи', 'count')
                ).reset_index()
            
            # --- ИСПРАВЛЕНИЕ: Передаем все 5 аргументов
            self.rebuild_dashboard(prod_stats, cat_stats, cat2_stats, has_profit, total_sales)
            
            self.build_side_dashboard() # Обновляем боковой дашборд тоже


if __name__ == "__main__":
    root = tk.Tk()
    app = SalesAnalyzerApp(root)
    root.state('zoomed')
    root.mainloop()